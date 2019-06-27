# My current work is to model novel light paths using multiple light sources combined to a single output.
# TracePro, a ray tracing software, uses a Monte Carlo analysis to show the probabilistic behavior of light
# in a given modeled light path.
# The goal of this script is to construct an excel spreadsheet showing the independent/dependent variables
# of a given TracePro ray tracing analysis (data comes in the form of a .txt file) with an end product
# being graphs of irradiance transmission vs. a selected independent variable.

# End goals for this script would be to provide user inputs to select how many independent variables there are,
# the respective annotation/values/incremental changes and potentially have the script graph for the user.

# For the second iteration, I have made variables user inputs to construct lists of variables and create loops for
# setting up the spreadsheet. Next phase will be to use loops to set up the spreadsheet with the indep. variables
# prior to transferrring the dep. variables from the .txt file.
# Will also need to include some printed statements to help guide the user and figure out how exactly they want the
# spreadsheet set up.

# Other uses to potentially be included are: to find the optimal results of the ray tracing data
# (irradiance % transmission), ...
from math import pi
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from itertools import chain
import string

wb = Workbook()
ws = wb.active
ws.title = 'Test Sheet'

# This sets up all the variables (dependent and independent) based on the users input.
# There can be multiple independent variables and it will convert them into lists.
dep_var = input('What is your dependent variable and associated units of measurement (if any): ')
num_of_var = int(input('How many independent variables were observed: '))
var_naming = []
var_unit = []
var_start_val = []
var_end_val = []
var_end_val2 = []
var_step_val = []
var_step_val2 = []
var_ranges = []
for i in range(1,num_of_var + 1):
    var_naming.append('variable' + str(i))
    var_unit.append('unit' + str(i))
    var_start_val.append('startingvalue' + str(i))
    var_end_val.append('endingvalue' + str(i))
    var_end_val2.append('2ndendingvalue' + str(i))
    var_step_val.append('stepvalue' + str(i))
    var_step_val2.append('2ndstepvalue' + str(i))
    var_ranges.append('variablerange' + str(i))

var_index = 0
while var_index <= num_of_var - 1:
    var_naming[var_index] = input('What is your #' + str(var_index + 1) + ' variable name?: ')
    var_unit[var_index] = input('What is your #' + str(var_index + 1) + ' variable unit of measurement?: ')
    var_start_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable starting loop value?: '))
    num_of_steps = input('Will you have two different ranges in a single loop? (Yes/No): ')
    if num_of_steps == 'Yes':
        var_step_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable 1st step value?: '))
        var_end_val[var_index] = int(
            input('What is your #' + str(var_index + 1) + ' variable 1st ending loop value?: '))
        var_step_val2[var_index] = int(input('What is your #' + str(var_index + 1) +
                                         ' variable 2nd step value?: '))
        var_end_val2[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable 2nd ending loop value?: '))
    else:
        var_step_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable step loop value?: '))
        var_end_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable ending loop value?: '))
    var_index += 1

print(var_naming)
print(var_unit)
print(var_start_val)
print(var_end_val)
print(var_end_val2)
print(var_step_val)
print(var_step_val2)

# This section will combine two different size loops of the same variable if answered yes above.
# An example would be varying the length of something and you go from 10-100 mm every 10 mm
# then 100-500 mm every 100 mm. This would combine them into a single list.
for i in range(0, num_of_var):
    num_of_steps = input('Will your #' + str(i + 1) + ' variable have two different step values? (Yes/No): ')
    if num_of_steps == 'No':
        var_ranges[i] = range(var_start_val[i], var_end_val[i] + var_step_val[i], var_step_val[i])
    else:
        var_ranges[i] = list(chain(range(var_start_val[i], var_end_val[i], var_step_val[i]),
                          range(var_end_val[i], var_end_val2[i] + var_step_val2[i], var_step_val2[i])))
    for j in var_ranges[i]:
        print(j)

print(var_ranges[1])

# These two equations were swiped from the internet to convert column letters ('A') to numbers ('1') for the respective
# inputs and vice versa
# Column letter to number
def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

# Column number to letter
def convertToTitle(num):
    title = ''
    alist = string.ascii_uppercase
    while num:
        mod = (num-1) % 26
        num = int((num - mod) / 26)
        title += alist[mod]
    return title[::-1]

# The origin position of where all other cells are based.
starting_row = 2
starting_col = 'B'

# This if loop will construct all the variables (except for the dependent text data) depending on how many variable
# inputs we have.
if num_of_var == 1: # One variable
    print(col2num(starting_col))
    ws[starting_col + str(starting_row)] = var_naming[num_of_var - 1] + ' (' + var_unit[num_of_var - 1] + ')'
    ws[convertToTitle(col2num(starting_col) + 1) + str(starting_row)] = dep_var
    num_of_loop_var = len(var_ranges[num_of_var - 1])
    print(num_of_loop_var)
    ind_var_index = starting_row + 1
    for i in var_ranges[num_of_var - 1]:
        ws.cell(row=ind_var_index, column=col2num(starting_col)).value = i
        ind_var_index += 1

elif num_of_var == 2: # Two variables
    # First independent variable
    num_of_loop_var = len(var_ranges[num_of_var - 2])
    print(num_of_loop_var)
    col_spacing_index = col2num(starting_col)
    print(col_spacing_index)
    row_spacing_index = starting_row + 2
    print(row_spacing_index)
    for i in var_ranges[num_of_var - 2]:
        ws.cell(row=starting_row, column=col_spacing_index).value = var_naming[num_of_var - 2]+ ' (' + var_unit[
            num_of_var - 2] + ')'
        ws.cell(row=starting_row, column=col_spacing_index + 1).value = i
        ws.column_dimensions[convertToTitle(col_spacing_index)].width = 13
        ws.column_dimensions[convertToTitle(col_spacing_index + 1)].width = 13
        # Second independent variable
        ws.cell(row=starting_row + 1, column=col_spacing_index).value = var_naming[num_of_var - 1] + ' (' + var_unit[
            num_of_var - 1] + ')'
        ws.cell(row=starting_row + 1, column=col_spacing_index + 1).value = dep_var
        for j in var_ranges[num_of_var - 1]:
            ws.cell(row=row_spacing_index, column=col_spacing_index).value = j
            print(j)
            row_spacing_index += 1
        col_spacing_index += 3
        row_spacing_index = starting_row + 2
else: # Three variables
    num_of_loop_var = len(var_ranges[num_of_var - 2])
    print(num_of_loop_var)
    col_spacing_index = col2num(starting_col)
    print(col_spacing_index)
    row1_spacing_index = starting_row
    print(row1_spacing_index)
    row2_spacing_index = row1_spacing_index + 3
    print(row2_spacing_index)
    # First variable
    for i in var_ranges[num_of_var - 3]:
        ws.cell(row=row1_spacing_index, column=col_spacing_index).value = var_naming[num_of_var - 3] + ' (' + var_unit[
            num_of_var - 3] + ')'
        ws.cell(row=row1_spacing_index, column=col_spacing_index + 1).value = i
        # Second variable
        for j in var_ranges[num_of_var - 2]:
            ws.cell(row=row1_spacing_index + 1, column=col_spacing_index).value = var_naming[num_of_var - 2] + ' (' + var_unit[num_of_var - 2] + ')'
            ws.cell(row=row1_spacing_index + 1, column=col_spacing_index + 1).value = j
            # Third variable
            ws.cell(row=row1_spacing_index + 2, column=col_spacing_index).value = var_naming[num_of_var - 1] + ' (' + var_unit[num_of_var - 1] + ')'
            ws.cell(row=row1_spacing_index + 2, column=col_spacing_index + 1).value = dep_var
            for k in var_ranges[num_of_var - 1]:
                ws.cell(row=row2_spacing_index, column=col_spacing_index).value = k
                row2_spacing_index += 1
            ws.column_dimensions[convertToTitle(col_spacing_index)].width = 14
            ws.column_dimensions[convertToTitle(col_spacing_index + 1)].width = 13
            row2_spacing_index = row1_spacing_index + 3
            col_spacing_index += 3
        row1_spacing_index += num_of_loop_var + 4
        col_spacing_index = col2num(starting_col)

# irr_data = open(input('Copy the path for the desired data .txt file you want to input: '), 'r')
# if num_of_var == 1: # One variable
#     irr_index = 3
#     col_spacing_index = col2num(starting_col)
#     for point in irr_data:
#         ws.cell(row=irr_index, column=col_spacing_index + 1).value = float(point)
#         irr_index += 1
# elif num_of_var == 2: # Two variables
#     irr_index = 4
#     col_spacing_index = col2num(starting_col)
#     for i in var_ranges[num_of_var - 1]:
#         for point in irr_data:
#             ws.cell(row=irr_index, column=col_spacing_index + 1).value = float(point)
#             irr_index += 1
#         col_spacing_index += 3
# else: # Three variables
#     irr_index = 5
#     irr_index2 = 5
#     num_of_loop_var = len(var_ranges[num_of_var - 2])
#     col_spacing_index = col2num(starting_col)
#     for i in var_ranges[num_of_var - 1]: # First variable
#         for j in var_ranges[num_of_var - 2]: # Second variable
#             for point in irr_data: # Third variable
#                 ws.cell(row=irr_index, column=col_spacing_index + 1).value = float(point)
#                 irr_index += 1
#             col_spacing_index += 3
#         irr_index2 += num_of_loop_var
#
# irr_data.close()
#
# chart = ScatterChart(scatterStyle='smoothMarker')
# chart.x_axis.title = 'Radius of 2nd Curve (mm)'
# chart.y_axis.title = '% Transmission'
#
# xvalues = Reference(ws, min_col=3, min_row=4, max_row=22)
# yvalues = Reference(ws, min_col=5, min_row=4, max_row=22)
# series = Series(yvalues, xvalues)
# chart.series.append(series)
#
# ws.add_chart(chart, 'G3')

# https://openpyxl.readthedocs.io/en/stable/charts/scatter.html refer to this for next iteration with multiple lines

wb.save('TestWorkbook5.xlsx')
