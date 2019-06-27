# My current work is to model novel light paths using multiple light sources combined to a single output.
# TracePro, a ray tracing software, uses a Monte Carlo analysis to show the probabilistic behavior of light
# in a given modeled light path.
# The goal of this script is to construct an excel spreadsheet showing the independent/dependent variables
# of a given TracePro ray tracing analysis (data comes in the form of a .txt file) with an end product
# being graphs of irradiance transmission vs. a selected independent variable.

# End goals for this script would be to provide user inputs to select how many independent variables there are,
# the respective annotation/values/incremental changes and potentially have the script graph for the user.

# For the second iteration, I have made variables user inputs to construct lists of variables and create loops for
# setting up the spreadsheet. Next phase will be to use loops to set up the spreadsheet with the indep. variables
# prior to transferrring the dep. variables from the .txt file.
# Will also need to include some printed statements to help guide the user and figure out how exactly they want the
# spreadsheet set up.

# Other uses to potentially be included are: to find the optimal results of the ray tracing data
# (irradiance % transmission), ...
from math import pi
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from itertools import chain

wb = Workbook()
ws = wb.active
ws.title = 'Test Sheet'

# This sets up all the variables (dependent and independent) based on the users input.
# There can be multiple independent variables and it will convert them into lists.
dep_var = input('What is your dependent variable and associated units of measurement (if any): ')
num_of_var = input('How many independent variables were observed: ')
var_naming = []
var_unit = []
var_start_val = []
var_end_val = []
var_end_val2 = []
var_step_val = []
var_step_val2 = []
var_ranges = []
for i in range(1,int(num_of_var) + 1):
    var_naming.append('variable' + str(i))
    var_unit.append('unit' + str(i))
    var_start_val.append('startingvalue' + str(i))
    var_end_val.append('endingvalue' + str(i))
    var_end_val2.append('2ndendingvalue' + str(i))
    var_step_val.append('stepvalue' + str(i))
    var_step_val2.append('2ndstepvalue' + str(i))
    var_ranges.append('variablerange' + str(i))
# Need to figure out how to account for the second step values. Maybe include an if statement in the below while loop with a second step value list.

var_index = 0
while var_index <= int(num_of_var) - 1:
    var_naming[var_index] = input('What is your #' + str(var_index + 1) + ' variable name?: ')
    var_unit[var_index] = input('What is your #' + str(var_index + 1) + ' variable unit of measurement?: ')
    var_start_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable starting loop value?: '))
    num_of_steps = input('Will you have two different ranges in a single loop? (Yes/No): ')
    if num_of_steps == 'Yes':
        var_step_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable 1st step value?: '))
        var_step_val2[var_index] = int(input('What is your #' + str(var_index + 1) +
                                         ' variable 2nd step value?: '))
        var_end_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable 1st ending loop value?: '))
        var_end_val2[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable 2nd ending loop value?: '))
    else:
        var_step_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable step loop value?: '))
        var_end_val[var_index] = int(input('What is your #' + str(var_index + 1) + ' variable ending loop value?: '))
    var_index += 1

print(var_naming)
print(var_unit)
print(var_start_val)
print(var_end_val)
print(var_end_val2)
print(var_step_val)
print(var_step_val2)

# This section will combine two different size loops of the same variable if answered yes above.
# An example would be varying the length of something and you go from 10-100 mm every 10 mm
# then 100-500 mm every 100 mm. This would combine them into a single list.
for i in range(0, int(num_of_var)):
    num_of_steps = input('Will your #' + str(i + 1) + ' have two different step values? (Yes/No): ')
    if num_of_steps == 'No':
        var_ranges[i] = range(var_start_val[i], var_end_val[i] + var_step_val[i], var_step_val[i])
    else:
        var_ranges[i] = chain(range(var_start_val[i], var_end_val[i] + var_step_val[i], var_step_val[i]),
                          range(var_end_val[i], var_end_val2[i] + var_step_val2[i], var_step_val2[i]))
    for j in var_ranges[i]:
        print(j)

print(var_ranges)

# This was the beginning of inputting the users variables into Excel cells
ws['B2'] = var_naming[0] + ' (' + var_unit[0] + ')'
ws['C2'] = var_start_val[0]
# Will have to make into a loop for later versions

# From here down everything is the same as the last version. I was not focused on how the Excel file looked in this
# iteration but how the user inputs were manipulated by printing them to the shell.
ws['B3'] = 'Radius of 1st Curve (mm)'
ws['C3'] = 10

ws['B4'] = 'Radius of 2nd Curve (mm)'
ws['D3'] = 'Total Arc Length (mm)'
ws['E3'] = '% Transmission'

# Calculates the arc length of the first and second curve totaling them together
# This is specific to this test so this section and the respective math would
# change for a different set of variables and loops.

SecondRadii = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
rad_index = 4
num_var = len(SecondRadii)
print(num_var)
final_ID_pos = 'C' + str(num_var)
print(final_ID_pos)
ws.column_dimensions['B'].width = 22 # See if you can convert text in .py file to pixels for this dimension to automate
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 13
cell_range = 'C4', final_ID_pos

for SecondRadius in SecondRadii:
    ws.cell(row = rad_index, column = 3).value = SecondRadius
    tot_arc = ((ws['C2'].value / 360) * 2 * pi * ws['C3'].value) + ((ws['C2'].value / 360) * 2 * pi * SecondRadius) # Need to change for final version to input the excel formula into each cell
    ws.cell(row = rad_index, column = 4).value = tot_arc
    rad_index += 1

irr_data = open('Test Irradiance Data.txt', 'r')
irr_index = 4
for point in irr_data:
    ws.cell(row = irr_index, column = 5).value = float(point)
    irr_index += 1

irr_data.close()

chart = ScatterChart(scatterStyle='smoothMarker')
chart.x_axis.title = 'Radius of 2nd Curve (mm)'
chart.y_axis.title = '% Transmission'

xvalues = Reference(ws, min_col=3, min_row=4, max_row=22)
yvalues = Reference(ws, min_col=5, min_row=4, max_row=22)
series = Series(yvalues, xvalues)
chart.series.append(series)

ws.add_chart(chart, 'G3')

# https://openpyxl.readthedocs.io/en/stable/charts/scatter.html refer to this for next iteration with multiple lines



wb.save('TestWorkbook2.xlsx')
