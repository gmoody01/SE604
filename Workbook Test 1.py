# My current work is to model novel light paths using multiple light sources combined to a single output.
# TracePro, a ray tracing software, uses a Monte Carlo analysis to show the probabilistic behavior of light
# in a given modeled light path.
# The goal of this script is to construct an excel spreadsheet showing the independent/dependent variables
# of a given TracePro ray tracing analysis (data comes in the form of a .txt file) with an end product
# being graphs of irradiance transmission vs. a selected independent variable.

# End goals for this script would be to provide user inputs to select how many independent variables there are,
# the respective annotation/values/incremental changes and potentially have the script graph for the user.

# For the first iteration, I am hard coding most variables to better understand the ext. module openpyxl
# and how to set up a spreadsheet with the code. Later, the user inputs will be included for a generalized use.

# Other uses to potentially be included are: to find the optimal results of the ray tracing data
# (irradiance % transmission), ...
from math import pi
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

# This sets up the workbook, worksheet and fills in the necessary title cells
wb = Workbook()
ws = wb.active
ws.title = 'Test Sheet'

ws['B2'] = 'Angle (degrees)'
ws['C2'] = 15

ws['B3'] = 'Radius of 1st Curve (mm)'
ws['C3'] = 10

ws['B4'] = 'Radius of 2nd Curve (mm)'
ws['D3'] = 'Total Arc Length (mm)'
ws['E3'] = '% Transmission'

# Calculates the arc length of the first and second curve totaling them together
# This is specific to this test so this section and the respective math would
# change for a different set of variables and loops.

# Here I listed out the independent variable data and altered row/column sizes
SecondRadii = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
rad_index = 4
num_var = len(SecondRadii)
print(num_var)
final_ID_pos = 'C' + str(num_var)
print(final_ID_pos)
ws.column_dimensions['B'].width = 22 # See if you can convert text in .py file to pixels for this dimension to automate
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 13
cell_range = 'C4', final_ID_pos

# This copies the independent variable data to the respective cells
for SecondRadius in SecondRadii:
    ws.cell(row = rad_index, column = 3).value = SecondRadius
    tot_arc = ((ws['C2'].value / 360) * 2 * pi * ws['C3'].value) + ((ws['C2'].value / 360) * 2 * pi * SecondRadius) # Need to change for final version to input the excel formula into each cell
    ws.cell(row = rad_index, column = 4).value = tot_arc
    rad_index += 1

# The dependent variable data is transferred over from the text file
irr_data = open('Test Irradiance Data.txt', 'r')
irr_index = 4
for point in irr_data:
    ws.cell(row = irr_index, column = 5).value = float(point)
    irr_index += 1

irr_data.close()

# This graphs the dependent variable vs. the independent variable
chart = ScatterChart(scatterStyle='smoothMarker')
chart.x_axis.title = 'Radius of 2nd Curve (mm)'
chart.y_axis.title = '% Transmission'

xvalues = Reference(ws, min_col=3, min_row=4, max_row=22)
yvalues = Reference(ws, min_col=5, min_row=4, max_row=22)
series = Series(yvalues, xvalues)
chart.series.append(series)

ws.add_chart(chart, 'G3')

# https://openpyxl.readthedocs.io/en/stable/charts/scatter.html refer to this for next iteration with multiple lines



wb.save('TestWorkbook.xlsx')
